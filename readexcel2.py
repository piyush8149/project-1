import openpyxl

wb=openpyxl.load_workbook("Book1.xlsx")
sh1=wb['Sheet1']
row=sh1.max_row
column=sh1.max_column
print(row)
print(column)


for i in range(1,row+1):
    for j in range(1,column+1):
        print(sh1.cell(i,j).value)


sh1.cell(row=7,column=2,value='pytest')
sh1.cell(row=7,column=3,value='UK')
sh1.cell(row=7,column=4,value='88.88')

wb.save("Report.xlsx")
