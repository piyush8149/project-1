import openpyxl

wb=openpyxl.load_workbook("Book1.xlsx")
sheets=wb.sheetnames
#print(wb.active.title)
sh1=wb['Sheet1']
data=sh1['B3'].value

#option1
print(wb['Sheet1']['c3'].value)

#option2
print(sh1.cell(5,3).value)
print(sh1.cell(5,4).value)

sh2=wb['Marks']
print(sh2.cell(2,2).value)

#option3
print(sh2.cell(row=2,column=2).value)


