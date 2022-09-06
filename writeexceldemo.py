from openpyxl import Workbook
from openpyxl.styles import PatternFill

# 7CFC00 green FF0000 red




wb=Workbook()
wb['Sheet']. title="Report of Automation"
sh1=wb.active
sh1['A1'].value="Name"
sh1['B1'].value="Status"
sh1['A2'].value="Python"
sh1['B2'].value="Active"
sh1['B2'].fill=PatternFill("solid",fgColor="71FF33")
sh1['A3'].value="Java"
sh1['B3'].value="Inactive"
sh1['B3'].fill=PatternFill("solid",fgColor="F50707")

wb.save("FinalReportNew.xlsx")